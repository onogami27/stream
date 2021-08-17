import openpyxl
import pandas as pd

def excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["金型修理改善依頼書"]
    #品番
    hinban = ws["C13"]
    #型番
    kataban = ws["C15"]
    #取数
    torisu = ws["F15"]
    #成型材質
    zaisitu = ws["I15"]
    #グレード
    grade = ws["I16"]
    #月産数
    gesan = ws["C17"]
    #製品形状
    keizyou = ws["I17"]
    #成型加工区
    kakouku = ws["M13"]
    #開始日
    kaishi = ws["M16"]
    #型納期
    nouki = ws["M17"]
    #製作メーカー
    seisaku = ws["Q15"]
    #修理メーカー
    syuuri = ws["Q16"]
    #修理区分
    kubun = ws["Q17"]
    #発行年月日
    hakou = ws["V13"]
    #手配ナンバー
    tehai_NO = ws["V15"]
    #ショットカウンタ
    shot_count = ws["Z22"]

    ind = ["品番","型番","取数","成形材質","材料グレード","月産数","製品形状","成形加工区","開始日",
        "型納期","製作メーカー","修理メーカー","修理区分","発行年月日","手配NO","ショットカウンタ数"]
    df = pd.DataFrame([[hinban.value],
                    [kataban.value],
                    [torisu.value],
                    [zaisitu.value],
                    [grade.value],
                    [gesan.value],
                    [keizyou.value],
                    [kakouku.value],
                    [str("{0:%Y/%m/%d}".format(kaishi.value))],
                    [str("{0:%Y/%m/%d}".format(nouki.value))],
                    [seisaku.value],
                    [syuuri.value],
                    [kubun.value],
                    [str("{0:%Y/%m/%d}".format(hakou.value))],
                    [tehai_NO.value],
                    [str(shot_count.value)]
                    ],index=ind,columns=["取得値"])
    return df


